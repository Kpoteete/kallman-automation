import socket
import ipaddress
import requests
from openpyxl import load_workbook, Workbook
from tkinter import Tk, filedialog
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime


TIMEOUT_SECONDS = 6
MAX_WORKERS = 20

# Helps prevent DNS lookups from hanging forever
socket.setdefaulttimeout(TIMEOUT_SECONDS)


def normalize_domain(value):
    if value is None:
        return None

    domain = str(value).strip()

    if not domain:
        return None

    domain = domain.replace(" ", "")

    if domain.startswith("http://") or domain.startswith("https://"):
        parsed = urlparse(domain)
        domain = parsed.netloc or parsed.path

    domain = domain.split("/")[0]

    # Remove common junk
    domain = domain.strip().strip(".").lower()

    return domain if domain else None


def website_is_good(domain):
    if not domain:
        return False, "Blank domain"

    urls_to_try = [
        f"https://{domain}",
        f"http://{domain}",
    ]

    headers = {
        "User-Agent": "Mozilla/5.0 WebsiteChecker/1.0"
    }

    last_error = None

    for url in urls_to_try:
        try:
            response = safe_get(url, headers)

            status_code = response.status_code
            response.close()

            # Count successful sites as Good
            if status_code < 400:
                return True, f"{status_code} OK"

            # Count 403 as Good because many valid sites block checkers/bots
            if status_code == 403:
                return True, "403 blocked checker"

            last_error = f"{status_code} error"

        except requests.exceptions.SSLError:
            last_error = "SSL error"

        except requests.exceptions.ConnectTimeout:
            last_error = "Connection timeout"

        except requests.exceptions.ReadTimeout:
            last_error = "Read timeout"

        except requests.exceptions.ConnectionError:
            last_error = "Connection error"

        except requests.RequestException as e:
            last_error = type(e).__name__

    return False, last_error or "Failed"


def validate_public_url(url):
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("Unsupported URL")
    if parsed.username or parsed.password or parsed.port not in {None, 80, 443}:
        raise ValueError("Credentials or nonstandard ports are not allowed")
    if parsed.hostname.lower() == "localhost":
        raise ValueError("Local addresses are not allowed")
    for _, _, _, _, sockaddr in socket.getaddrinfo(parsed.hostname, parsed.port or 443):
        address = ipaddress.ip_address(sockaddr[0])
        if not address.is_global:
            raise ValueError(f"Non-public address is not allowed: {address}")


def safe_get(url, headers, max_redirects=5):
    """GET only public HTTP(S) targets, validating every redirect hop."""
    from urllib.parse import urljoin
    current = url
    for _ in range(max_redirects + 1):
        validate_public_url(current)
        response = requests.get(
            current, headers=headers, timeout=TIMEOUT_SECONDS,
            allow_redirects=False, stream=True
        )
        if response.is_redirect or response.is_permanent_redirect:
            location = response.headers.get("Location")
            response.close()
            if not location:
                raise requests.RequestException("Redirect missing Location")
            current = urljoin(current, location)
            continue
        return response
    raise requests.TooManyRedirects("Too many redirects")


def check_domain_task(row_number, original_value):
    domain = normalize_domain(original_value)
    is_good, reason = website_is_good(domain)

    return {
        "row": row_number,
        "original_value": original_value,
        "domain": domain,
        "status": "Good" if is_good else "Bad",
        "reason": reason
    }


def select_excel_file():
    root = Tk()
    root.withdraw()

    return filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )


def save_results_to_excel(results, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    headers = [
        "Original Row",
        "Original Value",
        "Normalized Domain",
        "Status",
        "Reason"
    ]

    sheet.append(headers)

    for result in results:
        sheet.append([
            result["row"],
            result["original_value"],
            result["domain"],
            result["status"],
            result["reason"]
        ])

    workbook.save(output_path)


def main():
    file_path = select_excel_file()

    if not file_path:
        print("No file selected.")
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    jobs = []

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value

        if cell_value is None or str(cell_value).strip() == "":
            continue

        jobs.append((row, cell_value))

    total = len(jobs)

    print(f"Checking {total} domains using {MAX_WORKERS} workers...")
    print("This may take a few minutes depending on how many bad domains time out.")
    print()

    completed = 0
    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(check_domain_task, row_number, value): (row_number, value)
            for row_number, value in jobs
        }

        for future in as_completed(futures):
            completed += 1

            try:
                result = future.result()
            except Exception as e:
                row_number, value = futures[future]
                result = {
                    "row": row_number,
                    "original_value": value,
                    "domain": normalize_domain(value),
                    "status": "Bad",
                    "reason": f"Script error: {type(e).__name__}"
                }

            results.append(result)

            print(
                f"{completed}/{total} | "
                f"Row {result['row']} | "
                f"{result['domain']} | "
                f"{result['status']} | "
                f"{result['reason']}"
            )

    good_results = [
        result for result in results
        if result["status"] == "Good"
    ]

    bad_results = [
        result for result in results
        if result["status"] == "Bad"
    ]

    # Sort by original row number so the output follows the original Excel order
    good_results.sort(key=lambda x: x["row"] or 0)
    bad_results.sort(key=lambda x: x["row"] or 0)

    original_path = Path(file_path)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    good_output_path = original_path.with_name(
        f"{timestamp}_good.xlsx"
    )

    bad_output_path = original_path.with_name(
        f"{timestamp}_bad.xlsx"
    )

    save_results_to_excel(good_results, good_output_path)
    save_results_to_excel(bad_results, bad_output_path)

    print()
    print("Done.")
    print(f"Good domains saved here:")
    print(good_output_path)
    print()
    print(f"Bad domains saved here:")
    print(bad_output_path)


if __name__ == "__main__":
    main()
