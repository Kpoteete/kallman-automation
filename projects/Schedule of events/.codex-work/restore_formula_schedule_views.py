from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


PROJECT = Path(r"C:\kwi-automations\projects\Schedule of events")
SOURCE = PROJECT / "Schedule of events v - Codex weekly calendar view.xlsx"
OUTPUT = PROJECT / "Schedule of events v - formula schedule views restored.xlsx"

CATEGORY_COLORS = {
    "Forum": "FFDDEBFF",
    "Lunch": "FFFFF2CC",
    "Meet & Greet": "FFE2F0D9",
    "Opening Ceremony": "FFEADCF8",
    "Pavilion Tour": "FFDDEBF7",
    "Reception": "FFFCE4D6",
    "Summit / Keynote / Panel / Speaker": "FFE4DFEC",
    "Welcome Reception": "FFFCE4EC",
    "TPD1": "FFF2F2F2",
    "TPD2": "FFEAF7EA",
    "TPD3": "FFEAF2F8",
    "TPD4": "FFF8EAF2",
    "TPD5": "FFFFF7E6",
    "TPD6": "FFEDE7F6",
    "TPD7": "FFE0F2F1",
    "TPD8": "FFF3E5F5",
    "TPD9": "FFFBE9E7",
    "TPD10": "FFECEFF1",
}


def data_condition(date_ref: str, start_ref: str, end_ref: str, show_ref: str, slot_ref: str | None = None) -> str:
    parts = [
        f"(ROUND('Data Input'!$D$4:$D$484,10)<ROUND({end_ref},10))",
        f"(ROUND('Data Input'!$F$4:$F$484,10)>ROUND({start_ref},10))",
        f"('Data Input'!$A$4:$A$484={show_ref})",
        f"('Data Input'!$C$4:$C$484={date_ref})",
        "(($E$2=\"\")+('Data Input'!$B$4:$B$484=$E$2))",
        "(($F$2=\"\")+('Data Input'!$I$4:$I$484=$F$2))",
        "(($G$2=\"\")+('Data Input'!$J$4:$J$484=$G$2))",
        "(($H$2=\"\")+('Data Input'!$K$4:$K$484=$H$2))",
    ]
    if slot_ref:
        parts.insert(4, f"('Data Input'!$R$4:$R$484={slot_ref})")
    return "*".join(parts)


def daily_formula(row: int, col: int) -> str:
    slot = f"{get_column_letter(col)}$6"
    t = f"$A{row}"
    inc_end = f"$A{row}+TIME(0,$D$2,0)"
    cond = data_condition("$B$2", t, inc_end, "$C$2", slot)
    return (
        f'=IF({t}="","",IFERROR(LET('
        f'r,MATCH(1,{cond},0),'
        f'ev,INDEX(\'Data Input\'!$G$4:$G$484,r),'
        f'owner,INDEX(\'Data Input\'!$B$4:$B$484,r),'
        f'st,INDEX(\'Data Input\'!$D$4:$D$484,r),'
        f'en,INDEX(\'Data Input\'!$F$4:$F$484,r),'
        f'loc,INDEX(\'Data Input\'!$J$4:$J$484,r),'
        f'cls,INDEX(\'Data Input\'!$K$4:$K$484,r),'
        f'cat,INDEX(\'Data Input\'!$I$4:$I$484,r),'
        f'req,INDEX(\'Data Input\'!$L$4:$L$484,r),'
        f'IF((ROUND(st,10)>=ROUND({t},10))*(ROUND(st,10)<ROUND({inc_end},10)),'
        f'ev&CHAR(10)&CHAR(10)&IF(owner="","",owner&CHAR(10))&'
        f'TEXT(st,"h:mm AM/PM")&" - "&TEXT(en,"h:mm AM/PM")&CHAR(10)&'
        f'loc&CHAR(10)&CHAR(10)&'
        f'IF((cls="")*(cat=""),"",IF(cls="",cat,IF(cat="",cls,cls&" | "&cat)))&'
        f'IF(req="","",CHAR(10)&"Required: "&req)&CHAR(10),'
        f'IF(ROUND(en,10)<=ROUND({inc_end},10),".",""))'
        f'),""))'
    )


def weekly_condition(day_ref: str, start_ref: str, end_ref: str) -> str:
    return "*".join(
        [
            f"('Data Input'!$A$4:$A$484=$C$3)",
            f"('Data Input'!$C$4:$C$484={day_ref})",
            f"(ROUND('Data Input'!$D$4:$D$484,10)<ROUND({end_ref},10))",
            f"(ROUND('Data Input'!$F$4:$F$484,10)>ROUND({start_ref},10))",
            "(($D$3=\"\")+('Data Input'!$B$4:$B$484=$D$3))",
            "(($E$3=\"\")+('Data Input'!$I$4:$I$484=$E$3))",
            "(($F$3=\"\")+('Data Input'!$J$4:$J$484=$F$3))",
            "(($G$3=\"\")+('Data Input'!$K$4:$K$484=$G$3))",
        ]
    )


def weekly_formula(row: int, col: int) -> str:
    day = f"{get_column_letter(col)}$6"
    start = f"$A{row}"
    end = f"$A{row}+TIME(1,0,0)"
    cond = weekly_condition(day, start, end)
    return (
        f'=IFERROR(TEXTJOIN(CHAR(10)&CHAR(10),TRUE,FILTER('
        f'\'Data Input\'!$G$4:$G$484&CHAR(10)&'
        f'\'Data Input\'!$I$4:$I$484&CHAR(10)&'
        f'TEXT(\'Data Input\'!$D$4:$D$484,"h:mm AM/PM")&" - "&TEXT(\'Data Input\'!$F$4:$F$484,"h:mm AM/PM"),'
        f'{cond}'
        f')),"")'
    )


def add_category_cf(ws, target_range: str) -> None:
    ws.conditional_formatting._cf_rules.clear()
    for category, color in CATEGORY_COLORS.items():
        fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        escaped = category.replace('"', '""')
        ws.conditional_formatting.add(
            target_range,
            FormulaRule(formula=[f'ISNUMBER(SEARCH("{escaped}",B7))'], fill=fill),
        )


def main() -> None:
    copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)
    daily = wb["Daily Schedule"]
    weekly = wb["Weekly Calendar"]

    for row in range(7, 108):
        for col in range(2, 12):
            daily.cell(row=row, column=col).value = daily_formula(row, col)

    for row in range(7, 24):
        for col in range(2, 9):
            weekly.cell(row=row, column=col).value = weekly_formula(row, col)

    add_category_cf(daily, "B7:K107")
    add_category_cf(weekly, "B7:H23")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
