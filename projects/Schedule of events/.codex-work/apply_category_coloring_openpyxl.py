from __future__ import annotations

import shutil
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

project = Path(r"C:\kwi-automations\projects\Schedule of events")
source = project / "Schedule of events v - Codex weekly calendar view.xlsx"
output = project / "Schedule of events v - category color formatting.xlsx"
shutil.copy2(source, output)

colors = {
    "Forum": "DDEBFF",
    "Lunch": "FFF2CC",
    "Meet & Greet": "E2F0D9",
    "Opening Ceremony": "EADCF8",
    "Pavilion Tour": "DDEBF7",
    "Reception": "FCE4D6",
    "Summit / Keynote / Panel / Speaker": "E4DFEC",
    "Welcome Reception": "FCE4EC",
    "Keynote": "EAF3F8",
    "Speaker": "E4DFEC",
    "TPD1": "F2F2F2",
    "TPD2": "EAF7EA",
    "TPD3": "EAF2F8",
    "TPD4": "F8EAF2",
    "TPD5": "FFF7E6",
    "TPD6": "EDE7F6",
    "TPD7": "E0F2F1",
    "TPD8": "F3E5F5",
    "TPD9": "FBE9E7",
    "TPD10": "ECEFF1",
}

placeholders = [f"TPD{i}" for i in range(1, 11)]

def add_unique(ws, col: int, items: list[str], start_row: int = 2) -> None:
    existing = {
        str(ws.cell(r, col).value).strip()
        for r in range(start_row, 251)
        if ws.cell(r, col).value not in (None, "")
    }
    last = start_row - 1
    for r in range(start_row, 251):
        if ws.cell(r, col).value not in (None, ""):
            last = r
    for item in items:
        if item not in existing:
            last += 1
            ws.cell(last, col).value = item
            existing.add(item)

def serial_date(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return float((value - datetime(1899, 12, 30)).days)
    return float(value)

def serial_time(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400
    if isinstance(value, time):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400
    return float(value)

wb = load_workbook(output)
ref = wb["Reference"]
data = wb["Data Input"]
daily = wb["Daily Schedule"]
weekly = wb["Weekly Calendar"] if "Weekly Calendar" in wb.sheetnames else None

# Extend category lists and filter lists with placeholders.
add_unique(ref, 3, placeholders)
add_unique(ref, 10, placeholders, start_row=3)
add_unique(ref, 20, placeholders)

# Automatic category coloring on Daily Schedule.
daily.conditional_formatting._cf_rules.clear()
for category, color in colors.items():
    escaped = category.replace('"', '""')
    fill = PatternFill(fill_type="solid", fgColor=color)
    daily.conditional_formatting.add(
        "B7:K107",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("{escaped}",B7))'], fill=fill),
    )

# Static soft fills for the current Weekly Calendar cells based on underlying categories.
if weekly is not None:
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in weekly.iter_rows(min_row=7, max_row=23, min_col=2, max_col=8):
        for cell in row:
            cell.fill = white_fill

    week_start = serial_date(weekly["B3"].value)
    show_filter = str(weekly["C3"].value or "")
    owner_filter = str(weekly["D3"].value or "")
    category_filter = str(weekly["E3"].value or "")
    location_filter = str(weekly["F3"].value or "")
    class_filter = str(weekly["G3"].value or "")

    for day in range(7):
        for hour in range(6, 23):
            chosen_category = None
            for row in range(4, 485):
                show = data.cell(row, 1).value
                if not show:
                    continue
                if show_filter and show != show_filter:
                    continue
                owner = data.cell(row, 2).value or ""
                start_date = serial_date(data.cell(row, 3).value)
                start_time = serial_time(data.cell(row, 4).value)
                end_time = serial_time(data.cell(row, 6).value)
                event_name = data.cell(row, 7).value
                category = data.cell(row, 9).value or ""
                location = data.cell(row, 10).value or ""
                klass = data.cell(row, 11).value or ""
                if not event_name or start_date is None or start_time is None or end_time is None:
                    continue
                if owner_filter and owner != owner_filter:
                    continue
                if category_filter and category != category_filter:
                    continue
                if location_filter and location != location_filter:
                    continue
                if class_filter and klass != class_filter:
                    continue
                if int(start_date - week_start) != day:
                    continue
                if start_time < (hour + 1) / 24 and end_time > hour / 24:
                    chosen_category = category
                    break
            if chosen_category in colors:
                weekly.cell(7 + (hour - 6), 2 + day).fill = PatternFill(
                    fill_type="solid", fgColor=colors[chosen_category]
                )

wb.save(output)
print(output)
