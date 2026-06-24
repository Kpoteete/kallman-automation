from __future__ import annotations

from datetime import datetime, time, timedelta
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


PROJECT = Path(r"C:\kwi-automations\projects\Schedule of events")
SOURCE = PROJECT / "Schedule of events v - Codex weekly calendar view.xlsx"
OUTPUT = PROJECT / "Schedule of events v - daily weekly category colors.xlsx"

CATEGORY_COLORS = {
    "Opening Ceremony": "FFEADCF8",
    "Summit / Keynote / Panel / Speaker": "FFE4DFEC",
    "Lunch": "FFFFF2CC",
    "Pavilion Tour": "FFDDEBF7",
    "Forum": "FFDDEBFF",
    "Meet & Greet": "FFE2F0D9",
    "Reception": "FFFCE4D6",
    "Welcome Reception": "FFFCE4EC",
    "TPD1": "FFF2F2F2",
    "TPD2": "FFEAF7EA",
    "TPD3": "FFEAF2F8",
    "TPD4": "FFF8EAF2",
    "TPD5": "FFFFF7E6",
    "TPD6": "FFEDE7F6",
    "TPD7": "FFE0F2F1",
    "TPD8": "FFF3E5F5",
    "TPD9": "FFFBE9E7",
    "TPD10": "FFECEFF1",
}


def excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def excel_time(value):
    if isinstance(value, datetime):
        return value.time()
    return value


def fmt_time(value) -> str:
    value = excel_time(value)
    if isinstance(value, time):
        hour = value.hour
        minute = value.minute
        suffix = "AM" if hour < 12 else "PM"
        hour12 = hour % 12 or 12
        return f"{hour12}:{minute:02d} {suffix}"
    return str(value)


def add_tpd_categories(ref) -> None:
    for col in (3, 10, 20):
        existing = {
            str(ref.cell(row=row, column=col).value).strip()
            for row in range(1, ref.max_row + 1)
            if ref.cell(row=row, column=col).value not in (None, "")
        }
        row = 2
        for category in [f"TPD{i}" for i in range(1, 11)]:
            if category in existing:
                continue
            while isinstance(ref.cell(row=row, column=col), MergedCell) or ref.cell(row=row, column=col).value not in (None, ""):
                row += 1
            ref.cell(row=row, column=col).value = category
            existing.add(category)
            row += 1


def get_categories(ref) -> list[str]:
    categories: list[str] = []
    for row in range(2, ref.max_row + 1):
        value = ref.cell(row=row, column=3).value
        if value in (None, ""):
            continue
        text = str(value).strip()
        if not text or text.lower() == "categories":
            continue
        if text not in categories:
            categories.append(text)
    for category in CATEGORY_COLORS:
        if category not in categories:
            categories.append(category)
    return categories


def clear_cf(ws) -> None:
    ws.conditional_formatting._cf_rules.clear()


def add_contains_category_cf(ws, target_range: str, categories: list[str]) -> None:
    for category in categories:
        color = CATEGORY_COLORS.get(category, "FFF2F2F2")
        fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        escaped = category.replace('"', '""')
        formula = f'ISNUMBER(SEARCH("{escaped}",B7))'
        ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=fill))


def load_events(data):
    events = []
    for row in range(4, data.max_row + 1):
        show = data.cell(row=row, column=1).value
        start_date = excel_date(data.cell(row=row, column=3).value)
        start_time = excel_time(data.cell(row=row, column=4).value)
        end_time = excel_time(data.cell(row=row, column=6).value)
        name = data.cell(row=row, column=7).value
        category = data.cell(row=row, column=9).value
        owner = data.cell(row=row, column=2).value
        location = data.cell(row=row, column=10).value
        klass = data.cell(row=row, column=11).value
        if not (show and start_date and start_time and end_time and name):
            continue
        events.append(
            {
                "show": show,
                "owner": owner,
                "date": start_date,
                "start": start_time,
                "end": end_time,
                "name": str(name).strip().splitlines()[0].strip(),
                "category": str(category).strip() if category else "",
                "location": location,
                "class": klass,
            }
        )
    return events


def filter_matches(event, weekly) -> bool:
    show_filter = weekly["C3"].value
    owner_filter = weekly["D3"].value
    category_filter = weekly["E3"].value
    location_filter = weekly["F3"].value
    class_filter = weekly["G3"].value
    return (
        (not show_filter or event["show"] == show_filter)
        and (not owner_filter or event["owner"] == owner_filter)
        and (not category_filter or event["category"] == category_filter)
        and (not location_filter or event["location"] == location_filter)
        and (not class_filter or event["class"] == class_filter)
    )


def time_overlaps_hour(event, slot_start: time) -> bool:
    slot_dt = datetime.combine(datetime.today(), slot_start)
    slot_end = (slot_dt + timedelta(hours=1)).time()
    return event["start"] < slot_end and event["end"] > slot_start


def rebuild_weekly_text(weekly, events) -> None:
    for col in range(2, 9):
        day = excel_date(weekly.cell(row=6, column=col).value)
        for row in range(7, 24):
            slot = excel_time(weekly.cell(row=row, column=1).value)
            matches = [
                event
                for event in events
                if filter_matches(event, weekly)
                and event["date"] == day
                and time_overlaps_hour(event, slot)
            ]
            matches.sort(key=lambda e: (e["start"], e["name"]))
            blocks = []
            for event in matches:
                category_line = event["category"] if event["category"] else "Uncategorized"
                blocks.append(
                    f'{event["name"]}\n{category_line}\n{fmt_time(event["start"])} - {fmt_time(event["end"])}'
                )
            weekly.cell(row=row, column=col).value = "\n\n".join(blocks) if blocks else None


def main() -> None:
    copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)
    ref = wb["Reference"]
    add_tpd_categories(ref)
    categories = get_categories(ref)

    daily = wb["Daily Schedule"]
    clear_cf(daily)
    add_contains_category_cf(daily, "B7:K107", categories)

    weekly = wb["Weekly Calendar"]
    events = load_events(wb["Data Input"])
    rebuild_weekly_text(weekly, events)
    clear_cf(weekly)
    add_contains_category_cf(weekly, "B7:H23", categories)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
