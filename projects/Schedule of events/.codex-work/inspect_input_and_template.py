from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

project = Path(r"C:\kwi-automations\projects\Schedule of events")
input_path = project / "input.xlsx"
template_path = project / "Schedule of events v - Codex weekly calendar view.xlsx"

def cell_fill(cell):
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return None
    color = fill.fgColor
    return {
        "type": color.type,
        "rgb": color.rgb,
        "indexed": color.indexed,
        "theme": color.theme,
        "tint": color.tint,
    }

def preview_workbook(path: Path, max_rows=80, max_cols=14):
    wb = load_workbook(path, data_only=False)
    out = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        rows = []
        for r in range(1, min(ws.max_row, max_rows) + 1):
            vals = []
            nonempty = False
            for c in range(1, min(ws.max_column, max_cols) + 1):
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    nonempty = True
                vals.append(v.isoformat() if hasattr(v, "isoformat") else v)
            fills = []
            for c in range(1, min(ws.max_column, max_cols) + 1):
                f = cell_fill(ws.cell(r, c))
                fills.append(f)
            if nonempty:
                rows.append({"row": r, "values": vals, "fills": fills})
        out["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "preview_rows": rows[:max_rows],
        })
    return out

def template_summary(path: Path):
    wb = load_workbook(path, data_only=False)
    out = {"path": str(path), "sheets": wb.sheetnames}
    if "Data Input" in wb.sheetnames:
        ws = wb["Data Input"]
        out["data_input"] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "headers": [ws.cell(3, c).value for c in range(1, ws.max_column + 1)],
            "sample_rows": [
                [ws.cell(r, c).value.isoformat() if hasattr(ws.cell(r, c).value, "isoformat") else ws.cell(r, c).value for c in range(1, min(ws.max_column, 18) + 1)]
                for r in range(4, min(ws.max_row, 12) + 1)
            ],
        }
    if "Reference" in wb.sheetnames:
        ws = wb["Reference"]
        out["reference_preview"] = [
            [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
            for r in range(1, min(ws.max_row, 25) + 1)
        ]
    return out

result = {
    "input": preview_workbook(input_path),
    "template": template_summary(template_path),
}

print(json.dumps(result, indent=2, default=str))
