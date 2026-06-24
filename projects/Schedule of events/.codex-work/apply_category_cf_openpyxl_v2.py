from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

project = Path(r"C:\kwi-automations\projects\Schedule of events")
source = project / "Schedule of events v - Codex weekly calendar view.xlsx"
output = project / "Schedule of events v - category conditional formatting fixed.xlsx"
shutil.copy2(source, output)

colors = {
    "Forum": "DDEBFF",
    "Lunch": "FFF2CC",
    "Meet & Greet": "E2F0D9",
    "Opening Ceremony": "EADCF8",
    "Pavilion Tour": "DDEBF7",
    "Reception": "FCE4D6",
    "Summit / Keynote / Panel / Speaker": "E4DFEC",
    "Welcome Reception": "FCE4EC",
    "Keynote": "EAF3F8",
    "Speaker": "E4DFEC",
    "TPD1": "F2F2F2",
    "TPD2": "EAF7EA",
    "TPD3": "EAF2F8",
    "TPD4": "F8EAF2",
    "TPD5": "FFF7E6",
    "TPD6": "EDE7F6",
    "TPD7": "E0F2F1",
    "TPD8": "F3E5F5",
    "TPD9": "FBE9E7",
    "TPD10": "ECEFF1",
}
placeholders = [f"TPD{i}" for i in range(1, 11)]

def add_unique(ws, col: int, items: list[str], start_row: int = 2) -> None:
    existing = {
        str(ws.cell(r, col).value).strip()
        for r in range(start_row, 251)
        if ws.cell(r, col).value not in (None, "")
    }
    last = start_row - 1
    for r in range(start_row, 251):
        if ws.cell(r, col).value not in (None, ""):
            last = r
    for item in items:
        if item not in existing:
            last += 1
            ws.cell(last, col).value = item
            existing.add(item)

def dxf(color: str) -> DifferentialStyle:
    return DifferentialStyle(fill=PatternFill(fill_type="solid", fgColor=color))

wb = load_workbook(output)
ref = wb["Reference"]
daily = wb["Daily Schedule"]
weekly = wb["Weekly Calendar"] if "Weekly Calendar" in wb.sheetnames else None

add_unique(ref, 3, placeholders)
add_unique(ref, 10, placeholders, start_row=3)
add_unique(ref, 20, placeholders)

# Daily view: automatic text-based category coloring. These rules follow the selected day
# because the displayed event blocks recalculate from Data Input.
daily.conditional_formatting._cf_rules.clear()
for category, color in colors.items():
    rule = Rule(type="containsText", operator="containsText", text=category, dxf=dxf(color))
    rule.formula = [f'NOT(ISERROR(SEARCH("{category.replace(chr(34), chr(34)+chr(34))}",B7)))']
    daily.conditional_formatting.add("B7:K107", rule)

# Weekly calendar: automatic underlying-data coloring. The weekly cells only show title/time,
# so rules look back to Data Input by show/date/hour/category.
if weekly is not None:
    weekly.conditional_formatting._cf_rules.clear()
    for category, color in colors.items():
        escaped = category.replace('"', '""')
        formula = (
            "SUMPRODUCT("
            "('Data Input'!$A$4:$A$484=$C$3)*"
            "('Data Input'!$C$4:$C$484=B$6)*"
            "('Data Input'!$D$4:$D$484<($A7+TIME(1,0,0)))*"
            "('Data Input'!$F$4:$F$484>$A7)*"
            f"('Data Input'!$I$4:$I$484=\"{escaped}\")*"
            "(($D$3=\"\")+('Data Input'!$B$4:$B$484=$D$3))*"
            "(($E$3=\"\")+('Data Input'!$I$4:$I$484=$E$3))*"
            "(($F$3=\"\")+('Data Input'!$J$4:$J$484=$F$3))*"
            "(($G$3=\"\")+('Data Input'!$K$4:$K$484=$G$3))"
            ")>0"
        )
        weekly.conditional_formatting.add("B7:H23", FormulaRule(formula=[formula], fill=PatternFill(fill_type="solid", fgColor=color)))

wb.save(output)
print(output)
