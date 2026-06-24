from copy import copy
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


PROJECT = Path(r"C:\kwi-automations\projects\Schedule of events")
SOURCE = PROJECT / "Schedule of events v - Codex weekly calendar view.xlsx"
OUTPUT = PROJECT / "Schedule of events v - category conditional formatting fixed v2.xlsx"


CATEGORY_COLORS = {
    "Opening Ceremony": "EADCF8",
    "Summit / Keynote / Panel / Speaker": "E4DFEC",
    "Lunch": "FFF2CC",
    "Pavilion Tour": "DDEBF7",
    "Forum": "DDEBFF",
    "Meet & Greet": "E2F0D9",
    "Reception": "FCE4D6",
    "Welcome Reception": "FCE4EC",
    "TPD1": "F2F2F2",
    "TPD2": "EAF7EA",
    "TPD3": "EAF2F8",
    "TPD4": "F8EAF2",
    "TPD5": "FFF7E6",
    "TPD6": "EDE7F6",
    "TPD7": "E0F2F1",
    "TPD8": "F3E5F5",
    "TPD9": "FBE9E7",
    "TPD10": "ECEFF1",
}


def q(text: str) -> str:
    return text.replace('"', '""')


def add_unique_to_column(ws, col: int, values: list[str], start_row: int = 2) -> None:
    existing = {
        str(ws.cell(row=r, column=col).value).strip()
        for r in range(start_row, ws.max_row + 1)
        if ws.cell(row=r, column=col).value not in (None, "")
    }
    row = start_row
    while isinstance(ws.cell(row=row, column=col), MergedCell) or ws.cell(row=row, column=col).value not in (None, ""):
        row += 1
    for value in values:
        if value not in existing:
            while isinstance(ws.cell(row=row, column=col), MergedCell):
                row += 1
            ws.cell(row=row, column=col).value = value
            existing.add(value)
            row += 1


def clear_cf(ws) -> None:
    ws.conditional_formatting._cf_rules.clear()


def add_category_cf(ws, target_range: str, formula_template: str, categories: list[str]) -> None:
    for category in categories:
        color = CATEGORY_COLORS.get(category, "F2F2F2")
        fill = PatternFill(fill_type="solid", fgColor=color)
        formula = formula_template.format(category=q(category))
        ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=fill))


def build_weekly_helper(weekly) -> None:
    # AA:AG mirrors B:H and holds the category of the first event visible in each hour.
    helper_start_col = 27  # AA
    helper_end_col = helper_start_col + 6

    weekly.cell(row=1, column=helper_start_col).value = "Category helper - do not print"
    weekly.cell(row=5, column=helper_start_col).value = "Sunday"
    for offset in range(7):
        col = helper_start_col + offset
        source_day_col = 2 + offset
        weekly.cell(row=5, column=col).value = weekly.cell(row=5, column=source_day_col).value
        weekly.cell(row=6, column=col).value = f"={get_column_letter(source_day_col)}$6"

    for row in range(7, 24):
        for offset in range(7):
            col = helper_start_col + offset
            source_day = f"{get_column_letter(2 + offset)}$6"
            helper_cell = weekly.cell(row=row, column=col)
            helper_cell.value = (
                '=IFERROR(INDEX(\'Data Input\'!$I$4:$I$484,'
                'AGGREGATE(15,6,'
                '(ROW(\'Data Input\'!$I$4:$I$484)-ROW(\'Data Input\'!$I$4)+1)/'
                '(('
                f'\'Data Input\'!$A$4:$A$484=$C$3)*'
                f'(\'Data Input\'!$C$4:$C$484={source_day})*'
                f'(\'Data Input\'!$D$4:$D$484<($A{row}+TIME(1,0,0)))*'
                f'(\'Data Input\'!$F$4:$F$484>$A{row})*'
                '(($D$3="")+(\'Data Input\'!$B$4:$B$484=$D$3))*'
                '(($E$3="")+(\'Data Input\'!$I$4:$I$484=$E$3))*'
                '(($F$3="")+(\'Data Input\'!$J$4:$J$484=$F$3))*'
                '(($G$3="")+(\'Data Input\'!$K$4:$K$484=$G$3))'
                '),1)),"")'
            )
            helper_cell.font = copy(weekly.cell(row=row, column=2 + offset).font)

    for col in range(helper_start_col, helper_end_col + 1):
        letter = get_column_letter(col)
        weekly.column_dimensions[letter].hidden = True
        weekly.column_dimensions[letter].width = 14


def main() -> None:
    copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)

    ref = wb["Reference"]
    placeholders = [f"TPD{i}" for i in range(1, 11)]
    for col in (3, 10, 20):
        add_unique_to_column(ref, col, placeholders)

    categories = []
    for row in range(2, ref.max_row + 1):
        value = ref.cell(row=row, column=3).value
        if value not in (None, ""):
            text = str(value).strip()
            if text.lower() == "categories":
                continue
            if text and text not in categories:
                categories.append(text)
    for placeholder in placeholders:
        if placeholder not in categories:
            categories.append(placeholder)

    daily = wb["Daily Schedule"]
    clear_cf(daily)
    add_category_cf(
        daily,
        "B7:K107",
        'ISNUMBER(SEARCH("| {category}",B7))',
        categories,
    )

    weekly = wb["Weekly Calendar"]
    build_weekly_helper(weekly)
    clear_cf(weekly)
    add_category_cf(
        weekly,
        "B7:H23",
        'AA7="{category}"',
        categories,
    )

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
