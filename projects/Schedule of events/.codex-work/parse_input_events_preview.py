from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

path = Path(r"C:\kwi-automations\projects\Schedule of events\input.xlsx")
wb = load_workbook(path, data_only=True)
ws = wb["Schedule"]

for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v not in (None, "") for v in vals):
        print(r, vals)
