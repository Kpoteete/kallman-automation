from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def clean_code(value: object) -> str:
    text = clean(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(8) if digits else text


def latest_tally_workbook(root: Path) -> Path:
    pending = root / "01 Pending Review"
    matches = sorted(
        pending.glob("contact_interest_tally_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError(f"No contact_interest_tally_*.xlsx files found in {pending}")
    return matches[0]


def is_skip(value: str) -> bool:
    return value.strip().upper() in {"", "N/A", "NA", "SKIP", "NONE"}


def load_segment_mapping(path: Path) -> dict[str, tuple[str, str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    mapping: dict[str, tuple[str, str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        interest_code = clean(row[0]).upper()
        major = clean(row[1]).upper()
        minor = clean(row[2]).upper()
        combined = clean(row[3]).upper()
        if not interest_code or is_skip(major) or is_skip(minor) or is_skip(combined):
            continue
        mapping[interest_code] = (minor, major, combined)
    return mapping


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        width = min(60, max(len(str(cell.value or "")) for cell in column_cells) + 2)
        ws.column_dimensions[col_letter].width = max(10, width)


def main() -> int:
    parser = argparse.ArgumentParser(description="Map top contact-interest tally results to market segments.")
    parser.add_argument("--tally", default=None, help="Tally workbook. Defaults to latest 01 Pending Review/contact_interest_tally_*.xlsx.")
    parser.add_argument("--segments", default="Segments.xlsx", help="Interest-to-market-segment mapping workbook.")
    parser.add_argument("--output", default=None, help="Output workbook path.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    tally_path = Path(args.tally) if args.tally else latest_tally_workbook(root)
    if not tally_path.is_absolute():
        tally_path = root / tally_path

    segments_path = Path(args.segments)
    if not segments_path.is_absolute():
        segments_path = root / segments_path

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = Path(args.output) if args.output else root / "01 Pending Review" / f"market_segment_from_interest_tally_{timestamp}.xlsx"
    if not output_path.is_absolute():
        output_path = root / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    mapping = load_segment_mapping(segments_path)
    wb_tally = load_workbook(tally_path, read_only=True, data_only=True)
    ws_top = wb_tally["Top 3 By Account"]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Market Segment Output"
    ws_out.append([
        "AccountCode",
        "MarketSegmentMinorCode",
        "MarketSegmentMajorCode",
        "MarketSegmentMinorMajor",
        "MatchedRank",
        "MatchedInterestCode",
        "MatchedInterestName",
        "MatchedInterestCount",
        "TotalInterestVotes",
        "UniqueInterestCodes",
        "Message",
    ])

    matched = 0
    unmatched = 0
    for row in ws_top.iter_rows(min_row=2, values_only=True):
        account_code = clean_code(row[0])
        if not account_code:
            continue

        total_votes = row[3] or 0
        unique_codes = row[4] or 0
        candidates = [
            (1, clean(row[5]).upper(), clean(row[6]), row[7] or 0),
            (2, clean(row[8]).upper(), clean(row[9]), row[10] or 0),
            (3, clean(row[11]).upper(), clean(row[12]), row[13] or 0),
        ]

        selected = None
        for rank, interest_code, interest_name, count in candidates:
            if interest_code in mapping:
                selected = (rank, interest_code, interest_name, count, mapping[interest_code])
                break

        if selected is None:
            unmatched += 1
            ws_out.append([
                account_code,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                total_votes,
                unique_codes,
                "No mapped interest found in top 3.",
            ])
            continue

        matched += 1
        rank, interest_code, interest_name, count, segment = selected
        minor, major, combined = segment
        ws_out.append([
            account_code,
            minor,
            major,
            combined,
            rank,
            interest_code,
            interest_name,
            count,
            total_votes,
            unique_codes,
            "",
        ])

    ws_summary = wb_out.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Tally file", str(tally_path)])
    ws_summary.append(["Segments file", str(segments_path)])
    ws_summary.append(["Mapped segment codes loaded", len(mapping)])
    ws_summary.append(["Accounts matched", matched])
    ws_summary.append(["Accounts unmatched", unmatched])
    ws_summary.append(["Total output rows", matched + unmatched])

    for ws in wb_out.worksheets:
        style_sheet(ws)

    wb_out.save(output_path)
    print(f"Created: {output_path}")
    print(f"Accounts matched: {matched:,}")
    print(f"Accounts unmatched: {unmatched:,}")
    print(f"Mappings loaded: {len(mapping):,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
