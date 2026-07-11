from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INTEREST_CODE_RE = re.compile(r"\(([^()]+)\)")


def clean_code(value: object) -> str:
    text = "" if value is None else str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(8) if digits else text


def parse_interest_codes(raw: object) -> list[tuple[str, str]]:
    text = "" if raw is None else str(raw).strip()
    if not text:
        return []

    results: list[tuple[str, str]] = []
    seen: set[str] = set()
    for part in [piece.strip() for piece in text.split(",") if piece.strip()]:
        matches = INTEREST_CODE_RE.findall(part)
        for match in matches:
            code = match.strip().upper()
            if not code or code in seen:
                continue

            name = INTEREST_CODE_RE.sub("", part).strip(" -")
            results.append((code, name))
            seen.add(code)

    return results


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    max_width = 55
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        width = min(max_width, max(len(str(cell.value or "")) for cell in column_cells) + 2)
        ws.column_dimensions[col_letter].width = max(10, width)


def main() -> int:
    parser = argparse.ArgumentParser(description="Tally contact interest codes by primary account.")
    parser.add_argument(
        "input",
        nargs="?",
        default="contact with interest.xlsx",
        help="Input workbook. Defaults to 'contact with interest.xlsx' in this folder.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output workbook path. Defaults to 01 Pending Review/contact_interest_tally_<timestamp>.xlsx",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = root / input_path

    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = Path(args.output) if args.output else root / "01 Pending Review" / f"contact_interest_tally_{timestamp}.xlsx"
    if not output_path.is_absolute():
        output_path = root / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_in = load_workbook(input_path, read_only=True, data_only=True)
    ws_in = wb_in.active

    account_interest_counts: dict[str, Counter[str]] = defaultdict(Counter)
    account_interest_names: dict[str, dict[str, str]] = defaultdict(dict)
    account_interest_contacts: dict[tuple[str, str], set[str]] = defaultdict(set)
    account_rows: Counter[str] = Counter()
    account_market_segment: dict[str, str] = {}
    issue_rows: list[tuple[int, str, str, str]] = []

    data_rows = 0
    interest_votes = 0
    for row_number, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
        primary = clean_code(row[0] if len(row) > 0 else None)
        contact = clean_code(row[1] if len(row) > 1 else None)
        raw_interests = row[2] if len(row) > 2 else None
        market_segment = "" if len(row) <= 3 or row[3] is None else str(row[3]).strip()

        if not primary:
            issue_rows.append((row_number, "", contact, "Blank primary account code"))
            continue

        data_rows += 1
        account_rows[primary] += 1
        if market_segment and primary not in account_market_segment:
            account_market_segment[primary] = market_segment

        parsed = parse_interest_codes(raw_interests)
        if not parsed:
            issue_rows.append((row_number, primary, contact, "No interest code found in column C"))
            continue

        for interest_code, interest_name in parsed:
            account_interest_counts[primary][interest_code] += 1
            account_interest_names[primary].setdefault(interest_code, interest_name)
            if contact:
                account_interest_contacts[(primary, interest_code)].add(contact)
            interest_votes += 1

    wb_out = Workbook()
    ws_tally = wb_out.active
    ws_tally.title = "Interest Tally"
    ws_tally.append([
        "PrimaryAccountCode",
        "InterestCode",
        "InterestName",
        "Count",
        "ContactAccountCount",
        "ContactAccountCodes",
    ])

    for primary in sorted(account_interest_counts):
        for interest_code, count in sorted(
            account_interest_counts[primary].items(),
            key=lambda item: (-item[1], item[0]),
        ):
            contacts = sorted(account_interest_contacts[(primary, interest_code)])
            ws_tally.append([
                primary,
                interest_code,
                account_interest_names[primary].get(interest_code, ""),
                count,
                len(contacts),
                ", ".join(contacts),
            ])

    ws_top = wb_out.create_sheet("Top 3 By Account")
    ws_top.append([
        "PrimaryAccountCode",
        "SourceRows",
        "CurrentMarketSegment",
        "TotalInterestVotes",
        "UniqueInterestCodes",
        "Top1InterestCode",
        "Top1InterestName",
        "Top1Count",
        "Top2InterestCode",
        "Top2InterestName",
        "Top2Count",
        "Top3InterestCode",
        "Top3InterestName",
        "Top3Count",
    ])

    for primary in sorted(account_rows):
        counts = account_interest_counts.get(primary, Counter())
        top = sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:3]
        row = [
            primary,
            account_rows[primary],
            account_market_segment.get(primary, ""),
            sum(counts.values()),
            len(counts),
        ]
        for interest_code, count in top:
            row.extend([interest_code, account_interest_names[primary].get(interest_code, ""), count])
        while len(row) < 14:
            row.extend(["", "", 0])
        ws_top.append(row[:14])

    ws_summary = wb_out.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Input file", str(input_path)])
    ws_summary.append(["Source data rows", data_rows])
    ws_summary.append(["Primary accounts", len(account_rows)])
    ws_summary.append(["Interest tally rows", ws_tally.max_row - 1])
    ws_summary.append(["Total interest votes", interest_votes])
    ws_summary.append(["Rows with issues", len(issue_rows)])

    ws_issues = wb_out.create_sheet("Source Issues")
    ws_issues.append(["SourceRow", "PrimaryAccountCode", "ContactAccountCode", "Issue"])
    for issue in issue_rows:
        ws_issues.append(list(issue))

    for ws in wb_out.worksheets:
        style_sheet(ws)

    wb_out.save(output_path)
    print(f"Created: {output_path}")
    print(f"Primary accounts: {len(account_rows):,}")
    print(f"Interest tally rows: {ws_tally.max_row - 1:,}")
    print(f"Total interest votes: {interest_votes:,}")
    print(f"Rows with issues: {len(issue_rows):,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
