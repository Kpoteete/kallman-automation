from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def latest_segment_output(root: Path) -> Path:
    pending = root / "01 Pending Review"
    matches = sorted(
        pending.glob("market_segment_from_interest_tally_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError(f"No market_segment_from_interest_tally_*.xlsx files found in {pending}")
    return matches[0]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        width = min(70, max(len(str(cell.value or "")) for cell in column_cells) + 2)
        ws.column_dimensions[col_letter].width = max(10, width)


def main() -> int:
    parser = argparse.ArgumentParser(description="Create an approved Momentus market-segment apply workbook.")
    parser.add_argument("--input", default=None, help="Market segment output workbook. Defaults to latest pending review output.")
    parser.add_argument("--output", default=None, help="Approved apply workbook path.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    input_path = Path(args.input) if args.input else latest_segment_output(root)
    if not input_path.is_absolute():
        input_path = root / input_path

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = Path(args.output) if args.output else root / "02 Approved" / f"market_segment_apply_from_interest_tally_{timestamp}.xlsx"
    if not output_path.is_absolute():
        output_path = root / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_in = load_workbook(input_path, read_only=True, data_only=True)
    ws_in = wb_in["Market Segment Output"]

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Actions"
    headers = [
        "ReviewDecision",
        "Action",
        "Status",
        "ParentAccountCode",
        "ParentAccountName",
        "ContactAccountCode",
        "ContactName",
        "ContactEmail",
        "InterestCodes",
        "MarketSegmentMajor",
        "MarketSegmentMinor",
        "MarketSegmentCombined",
        "TargetSegmentAccountCode",
        "Message",
    ]
    ws.append(headers)

    approved = 0
    skipped_unmatched = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        account_code = clean(row[0])
        minor = clean(row[1])
        major = clean(row[2])
        combined = clean(row[3])
        matched_rank = clean(row[4])
        interest_code = clean(row[5])
        interest_name = clean(row[6])
        interest_count = clean(row[7])
        message = clean(row[10])

        if not account_code:
            continue

        if message or not (minor and major and combined):
            skipped_unmatched += 1
            continue

        approved += 1
        ws.append([
            "APPROVED",
            "UpdateParentMarketSegment",
            "PROPOSED",
            account_code,
            "",
            "",
            "",
            "",
            interest_code,
            major,
            minor,
            combined,
            account_code,
            f"Matched top-{matched_rank} interest {interest_code} ({interest_name}) with {interest_count} vote(s).",
        ])

    summary = wb_out.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Source file", str(input_path)])
    summary.append(["Approved update rows", approved])
    summary.append(["Skipped unmatched rows", skipped_unmatched])

    for sheet in wb_out.worksheets:
        style_sheet(sheet)

    wb_out.save(output_path)
    print(f"Created: {output_path}")
    print(f"Approved update rows: {approved:,}")
    print(f"Skipped unmatched rows: {skipped_unmatched:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
