from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - startup guidance
    raise SystemExit("Install dependencies first: py -m pip install -r requirements.txt") from exc


DEFAULT_URL = "https://kallman.ungerboeck.com/prod/app85.cshtml"
CODE_RE = re.compile(r"\b\d{4,}\b")


@dataclass(frozen=True)
class Pair:
    row: int
    merge_from: str
    merge_into: str

    @property
    def pair_id(self) -> str:
        raw = f"{self.row}|{self.merge_from}|{self.merge_into}"
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:20]


def clean_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def code_text(value: object, width: int) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value).zfill(width)
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).zfill(width)
    return str(value).strip()


def find_columns(headers: Iterable[object]) -> tuple[int, int]:
    normalized = [clean_header(value) for value in headers]
    aliases = {
        "from": {"merge from", "merge_from", "from", "account code from"},
        "into": {"merge into", "merge_into", "into", "account code into"},
    }
    from_idx = next((i for i, h in enumerate(normalized) if h in aliases["from"]), None)
    into_idx = next((i for i, h in enumerate(normalized) if h in aliases["into"]), None)
    if from_idx is None or into_idx is None:
        raise ValueError("The file must contain columns named 'Merge from' and 'Merge into'.")
    return from_idx, into_idx


def load_pairs(path: Path | str, width: int = 8) -> list[Pair]:
    path = Path(path)
    suffix = path.suffix.lower()
    rows: list[tuple[int, list[object]]]
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [(idx, list(row)) for idx, row in enumerate(sheet.iter_rows(values_only=True), 1)]
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [(idx, list(row)) for idx, row in enumerate(csv.reader(handle, delimiter=delimiter), 1)]
    else:
        raise ValueError("Supported input files are .xlsx, .xlsm, .csv, and .tsv.")

    if not rows:
        raise ValueError("The input file is empty.")
    from_idx, into_idx = find_columns(rows[0][1])
    pairs: list[Pair] = []
    seen: set[str] = set()
    for row_number, values in rows[1:]:
        merge_from = code_text(values[from_idx] if from_idx < len(values) else "", width)
        merge_into = code_text(values[into_idx] if into_idx < len(values) else "", width)
        if not merge_from and not merge_into:
            continue
        if not merge_from or not merge_into:
            raise ValueError(f"Row {row_number} has only one account code: {merge_from!r} -> {merge_into!r}")
        if merge_from == merge_into:
            raise ValueError(f"Row {row_number} merges an account into itself: {merge_from}")
        pair = Pair(row_number, merge_from, merge_into)
        if pair.pair_id in seen:
            raise ValueError(f"Duplicate merge pair on row {row_number}: {merge_from} -> {merge_into}")
        seen.add(pair.pair_id)
        pairs.append(pair)
    if not pairs:
        raise ValueError("No nonblank merge pairs were found.")
    return pairs


class ExcelStatusWriter:
    """Write each pair's latest program status back to the source Excel workbook."""

    STATUS_HEADER = "Merge Status"
    MESSAGE_HEADER = "Merge Message"
    UPDATED_HEADER = "Merge Updated At"

    def __init__(self, path: Path):
        self.path = Path(path)
        self.enabled = self.path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}
        self.workbook = None
        self.sheet = None
        self.status_col = None
        self.message_col = None
        self.updated_col = None
        if not self.enabled:
            print("Input is not an Excel workbook; statuses will remain in progress.csv only.")
            return

        keep_vba = self.path.suffix.lower() in {".xlsm", ".xltm"}
        self.workbook = load_workbook(self.path, keep_vba=keep_vba)
        self.sheet = self.workbook.active
        headers = {clean_header(cell.value): cell.column for cell in self.sheet[1]}
        self.status_col = headers.get(clean_header(self.STATUS_HEADER)) or self.sheet.max_column + 1
        self.message_col = headers.get(clean_header(self.MESSAGE_HEADER)) or max(self.sheet.max_column + 1, self.status_col + 1)
        self.updated_col = headers.get(clean_header(self.UPDATED_HEADER)) or max(self.sheet.max_column + 1, self.message_col + 1)

        self.sheet.cell(1, self.status_col, self.STATUS_HEADER)
        self.sheet.cell(1, self.message_col, self.MESSAGE_HEADER)
        self.sheet.cell(1, self.updated_col, self.UPDATED_HEADER)
        for col in (self.status_col, self.message_col, self.updated_col):
            self.sheet.cell(1, col).font = self.sheet.cell(1, 1).font.copy(bold=True)
            self.sheet.cell(1, col).fill = self.sheet.cell(1, 1).fill.copy()
            self.sheet.cell(1, col).border = self.sheet.cell(1, 1).border.copy()
            self.sheet.cell(1, col).alignment = self.sheet.cell(1, 1).alignment.copy()
        self.sheet.column_dimensions[self.sheet.cell(1, self.status_col).column_letter].width = 16
        self.sheet.column_dimensions[self.sheet.cell(1, self.message_col).column_letter].width = 45
        self.sheet.column_dimensions[self.sheet.cell(1, self.updated_col).column_letter].width = 22
        self.save()

    def write(self, pair: Pair, status: str, message: str = "") -> None:
        if not self.enabled:
            return
        display_status = {
            "processing": "Processing",
            "merged": "Merged",
            "skipped": "Skipped",
            "error": "Error",
            "human_needed": "Human Needed",
        }.get(status, status.title())
        updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(pair.row, self.status_col, display_status)
        self.sheet.cell(pair.row, self.message_col, message)
        self.sheet.cell(pair.row, self.updated_col, updated)
        self.sheet.cell(pair.row, self.updated_col).number_format = "yyyy-mm-dd hh:mm:ss"
        self.save()

    def sync_from_ledger(self, pairs: list[Pair], ledger: "Ledger") -> None:
        if not self.enabled:
            return
        for pair in pairs:
            row = ledger.db.execute(
                "select status,message,updated_at from pairs where pair_id = ?", (pair.pair_id,)
            ).fetchone()
            if not row:
                continue
            status, message, updated_at = row
            self.sheet.cell(pair.row, self.status_col, {
                "processing": "Processing", "merged": "Merged",
                "skipped": "Skipped", "error": "Error",
                "human_needed": "Human Needed"
            }.get(status, str(status).title()))
            self.sheet.cell(pair.row, self.message_col, message or "")
            try:
                parsed = datetime.fromisoformat(updated_at.replace("Z", "+00:00")).astimezone().replace(tzinfo=None)
                self.sheet.cell(pair.row, self.updated_col, parsed)
                self.sheet.cell(pair.row, self.updated_col).number_format = "yyyy-mm-dd hh:mm:ss"
            except Exception:
                self.sheet.cell(pair.row, self.updated_col, updated_at)
        self.save()

    def save(self) -> None:
        if not self.enabled:
            return
        try:
            self.workbook.save(self.path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Could not write status to {self.path}. Close the workbook in Excel and try again."
            ) from exc

    def close(self) -> None:
        if self.workbook is not None:
            self.workbook.close()


class Ledger:
    def __init__(self, path: Path):
        self.path = path
        self.db = sqlite3.connect(path)
        self.db.execute(
            """create table if not exists pairs (
                pair_id text primary key,
                row_number integer not null,
                merge_from text not null,
                merge_into text not null,
                status text not null,
                message text not null default '',
                updated_at text not null
            )"""
        )
        self.db.commit()

    def status(self, pair: Pair) -> str | None:
        row = self.db.execute("select status from pairs where pair_id = ?", (pair.pair_id,)).fetchone()
        return row[0] if row else None

    def record(self, pair: Pair, status: str, message: str = "") -> None:
        now = datetime.now(timezone.utc).isoformat()
        self.db.execute(
            """insert into pairs(pair_id,row_number,merge_from,merge_into,status,message,updated_at)
               values(?,?,?,?,?,?,?)
               on conflict(pair_id) do update set status=excluded.status,
               message=excluded.message, updated_at=excluded.updated_at""",
            (pair.pair_id, pair.row, pair.merge_from, pair.merge_into, status, message, now),
        )
        self.db.commit()

    def export_csv(self, path: Path) -> None:
        rows = self.db.execute(
            "select row_number,merge_from,merge_into,status,message,updated_at from pairs order by row_number"
        ).fetchall()
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["row", "merge_from", "merge_into", "status", "message", "updated_at"])
            writer.writerows(rows)

    def close(self) -> None:
        self.db.close()


def create_run_dir(root: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = root / stamp
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def visible_locator(locator, label: str):
    for index in range(locator.count()):
        candidate = locator.nth(index)
        if candidate.is_visible(timeout=1_000):
            return candidate
    raise RuntimeError(f"{label} is not visible")




def find_merge_into_field(page, merge_dialog, timeout_ms: int = 15000):
    """Find the Merge Into account lookup across known Momentus render variants."""
    deadline = time.monotonic() + (timeout_ms / 1000)
    last_details = []

    while time.monotonic() < deadline:
        # Momentus renders this lookup as a custom Aurelia combobox. In some
        # sessions the visible modal is not exposed as an ARIA dialog, so do
        # not rely only on merge_dialog. The fieldset text and CSS class are
        # much more stable in the captured production DOM.
        merge_into_fieldset = page.locator("fieldset").filter(has_text=re.compile(r"Merge Into Account", re.I))
        candidates = [
            ("Merge Into fieldset custom combobox", merge_into_fieldset.locator("input.ux-combobox-multiselect-input:visible")),
            ("Merge Into fieldset visible input", merge_into_fieldset.locator('input:visible:not([type="hidden"])')),
            ("page custom combobox", page.locator("input.ux-combobox-multiselect-input:visible")),
            ("dialog labeled textbox", merge_dialog.get_by_role("textbox", name=re.compile("merge.*into|account", re.I))),
            ("dialog combobox", merge_dialog.get_by_role("combobox")),
            ("dialog visible input", merge_dialog.locator('input:visible:not([type="hidden"])')),
            ("dialog visible textarea", merge_dialog.locator('textarea:visible')),
            ("page labeled textbox", page.get_by_role("textbox", name=re.compile("merge.*into|merge into account", re.I))),
            ("page labeled combobox", page.get_by_role("combobox", name=re.compile("merge.*into|merge into account", re.I))),
        ]

        for label, locator in candidates:
            try:
                count = locator.count()
                for i in range(count - 1, -1, -1):
                    candidate = locator.nth(i)
                    if candidate.is_visible(timeout=250) and candidate.is_enabled(timeout=250):
                        return candidate, label
            except Exception as exc:
                last_details.append(f"{label}: {exc}")

        page.wait_for_timeout(350)

    raise RuntimeError(
        "Merge Into account field was not found after waiting. "
        + (" | ".join(last_details[-3:]) if last_details else "No usable input candidates appeared.")
    )


def find_exact_merge_target(page, merge_dialog, account_code: str, timeout_ms: int = 15_000):
    """Find the visible autocomplete result whose displayed text contains the exact account code."""
    deadline = time.monotonic() + timeout_ms / 1000
    escaped = re.escape(account_code)
    patterns = [
        re.compile(rf"\({escaped}\)\s*$"),
        re.compile(rf"\b{escaped}\b"),
    ]
    while time.monotonic() < deadline:
        scopes = [merge_dialog, page]
        candidates = []
        for scope in scopes:
            for selector in (
                '[role="option"]:visible',
                '[role="button"]:visible',
                'button:visible',
                'li:visible',
                '.wj-listbox-item:visible',
                '.wj-listbox-item',
                '[class*="list-item"]:visible',
                '[class*="option"]:visible',
            ):
                loc = scope.locator(selector)
                for i in range(loc.count()):
                    item = loc.nth(i)
                    try:
                        if not item.is_visible(timeout=250):
                            continue
                        label = (item.inner_text(timeout=500) or '').strip()
                    except Exception:
                        continue
                    if label and any(p.search(label) for p in patterns):
                        candidates.append((item, label, selector))
        # Deduplicate by text and prefer the most exact-looking result.
        unique = []
        seen = set()
        for item, label, selector in candidates:
            key = (label, selector)
            if key not in seen:
                seen.add(key)
                unique.append((item, label, selector))
        exact = [x for x in unique if patterns[0].search(x[1])]
        chosen = exact if exact else unique
        if len(chosen) == 1:
            item, label, selector = chosen[0]
            return item, f"{selector}: {label}"
        if len(chosen) > 1:
            # Multiple DOM wrappers can represent the same visible option. Pick the smallest visible element.
            ranked = []
            for item, label, selector in chosen:
                try:
                    box = item.bounding_box() or {}
                    area = (box.get('width') or 99999) * (box.get('height') or 99999)
                except Exception:
                    area = 999999999
                ranked.append((area, item, label, selector))
            ranked.sort(key=lambda x: x[0])
            _, item, label, selector = ranked[0]
            return item, f"{selector}: {label}"
        page.wait_for_timeout(500)
    raise RuntimeError(f"Exact Merge Into account was not found in autocomplete: {account_code}")

def write_summary(path: Path, pairs: list[Pair], ledger: Ledger, input_path: Path, dry_run: bool) -> None:
    counts = {
        status: sum(1 for pair in pairs if ledger.status(pair) == status)
        for status in ("merged", "skipped", "human_needed", "error")
    }
    payload = {
        "input": str(input_path),
        "dry_run": dry_run,
        "total_pairs": len(pairs),
        "counts": counts,
        "remaining": len(pairs) - sum(counts.values()),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Resumable Momentus account merge UI runner")
    parser.add_argument("input", type=Path, help="Excel/CSV/TSV file with Merge from and Merge into columns")
    parser.add_argument("--url", default=DEFAULT_URL, help="Momentus Accounts URL")
    parser.add_argument("--runs-dir", type=Path, default=Path("runs"))
    parser.add_argument("--profile-dir", type=Path, default=Path(".momentus-profile"))
    parser.add_argument("--ledger", type=Path, help="Existing SQLite ledger to resume")
    parser.add_argument("--start-index", type=int, default=0, help="Zero-based pair index to begin processing")
    parser.add_argument("--limit", type=int, default=0, help="Process at most N pending pairs")
    parser.add_argument("--code-width", type=int, default=8)
    parser.add_argument("--dry-run", action="store_true", help="Search accounts but do not merge")
    parser.add_argument("--headed", action="store_true", default=True, help="Show the browser (default)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pairs = load_pairs(args.input, args.code_width)
    run_dir = args.ledger.parent if args.ledger else create_run_dir(args.runs_dir)
    ledger_path = args.ledger or run_dir / "progress.sqlite"
    ledger = Ledger(ledger_path)
    excel_status = ExcelStatusWriter(args.input)
    excel_status.sync_from_ledger(pairs, ledger)
    ledger.export_csv(run_dir / "progress.csv")
    write_summary(run_dir / "summary.json", pairs, ledger, args.input, args.dry_run)

    print(f"Loaded {len(pairs)} merge pairs from {args.input}")
    print(f"Run folder: {run_dir}")
    print("Use Ctrl+C to stop; progress is saved after every pair.")

    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        ledger.close()
        raise SystemExit("Install dependencies first: py -m pip install -r requirements.txt") from exc

    processed = 0
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(args.profile_dir),
            channel="chrome",
            headless=False,
            no_viewport=True,
            args=["--start-maximized"],
            ignore_default_args=["--no-sandbox"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(args.url, wait_until="commit", timeout=60_000)
        print("Browser opened. Complete Momentus login in the browser if prompted.")
        page.wait_for_timeout(5000)
        auto_user = os.environ.get("MOMENTUS_USER", "").strip()
        auto_password = os.environ.get("MOMENTUS_PASSWORD", "")
        login_user = page.get_by_role("textbox", name=re.compile("Email Address or User ID", re.I))
        if not login_user.count():
            try:
                login_user.wait_for(timeout=20_000)
            except PlaywrightTimeoutError:
                pass
        if login_user.count() and auto_user and auto_password:
            login_user.fill(auto_user)
            page.get_by_role("textbox", name=re.compile("Password", re.I)).fill(auto_password)
            page.get_by_role("button", name="Submit", exact=True).click()
            page.wait_for_timeout(2500)
        accounts_text = page.get_by_text("Accounts", exact=True).first
        deadline = time.monotonic() + 120
        while time.monotonic() < deadline:
            try:
                if accounts_text.is_visible(timeout=1_000):
                    break
            except PlaywrightTimeoutError:
                pass
            print("Waiting for Momentus Accounts. Complete login in the visible browser if prompted.")
            page.wait_for_timeout(2_000)
        else:
            raise RuntimeError("Momentus Accounts did not become available within 120 seconds")
        print(f"URL before Accounts navigation: {page.url}")
        nav_accounts = page.get_by_role("button", name="Accounts", exact=True).first
        if not nav_accounts.is_visible(timeout=1_000):
            raise RuntimeError("Accounts navigation button is not visible")
        nav_accounts.click()
        page.wait_for_timeout(3000)
        visible_locator(
            page.get_by_role("button", name="Apply Filters", exact=True),
            "Apply Filters button after navigation",
        )

        for pair in pairs[max(args.start_index, 0):]:
            if args.limit and processed >= args.limit:
                break
            if ledger.status(pair) in {"merged", "skipped", "human_needed"}:
                continue
            excel_status.write(pair, "processing", "Automation is working on this merge")
            try:
                clear_filters = visible_locator(page.get_by_role("button", name="Clear All", exact=True), "Clear All button")
                clear_filters.click()
                page.wait_for_timeout(1_200)
                account_inputs = page.locator('input[id^="424_"][id$="__control"]')
                if account_inputs.count():
                    account_code = visible_locator(account_inputs, "Account Code field")
                else:
                    account_code = visible_locator(
                        page.get_by_role("textbox", name=re.compile("Account Code", re.I)),
                        "Account Code field",
                    )
                account_code.click()
                account_code.press("Control+A")
                account_code.type(pair.merge_from, delay=20)
                print(f"ENTERED Account Code: {account_code.input_value()}")
                account_code.press("Tab")
                page.wait_for_timeout(400)
                visible_locator(page.get_by_role("button", name="Apply Filters", exact=True), "Apply Filters button").click()
                data_rows = []
                for _ in range(20):
                    rows = page.locator(".wj-row")
                    data_rows = []
                    for i in range(rows.count()):
                        candidate = rows.nth(i)
                        text = candidate.inner_text().strip()
                        if re.search(r"\b(?:Inactive|Active|Prospective)\b", text):
                            data_rows.append(candidate)
                    if data_rows:
                        break
                    page.wait_for_timeout(500)
                criteria = page.get_by_text(re.compile("Criteria:", re.I)).last
                criteria_text = criteria.inner_text().strip() if criteria.count() else ""
                print(f"CRITERIA after Apply: {criteria_text}")
                print(f"URL after Apply: {page.url}")
                page.screenshot(path=str(run_dir / f"after_apply_{pair.merge_from}.png"), full_page=False)
                (run_dir / f"after_apply_{pair.merge_from}.txt").write_text(
                    page.locator("body").inner_text(), encoding="utf-8"
                )
                print(f"CHECK {pair.merge_from}: {len(data_rows)} visible result rows")
                if not data_rows:
                    ledger.record(pair, "skipped", "No result for Merge From account code")
                    excel_status.write(pair, "skipped", "No result for Merge From account code")
                    print(f"SKIP {pair.merge_from} -> {pair.merge_into}: no result")
                    processed += 1
                    continue
                if len(data_rows) != 1:
                    raise RuntimeError(f"Expected one result, found {len(data_rows)}")

                if args.dry_run:
                    ledger.record(pair, "skipped", "Dry run: result found, merge not submitted")
                    excel_status.write(pair, "skipped", "Dry run: result found, merge not submitted")
                    print(f"DRY RUN {pair.merge_from} -> {pair.merge_into}")
                    processed += 1
                    continue

                row = data_rows[0]
                cell = row.locator('[role="gridcell"]').first
                try:
                    # The source account can disappear between the search result
                    # check and this click when a prior merge finishes in the
                    # background. Treat that as already completed and continue.
                    cell.click(timeout=5_000)
                    cell.click(button="right", timeout=5_000)
                except PlaywrightTimeoutError:
                    visible_rows_now = []
                    rows_now = page.locator(".wj-row")
                    for i in range(rows_now.count()):
                        candidate = rows_now.nth(i)
                        try:
                            text = candidate.inner_text(timeout=500).strip()
                        except Exception:
                            continue
                        if re.search(r"\b(?:Inactive|Active|Prospective)\b", text):
                            visible_rows_now.append(candidate)
                    no_rows_message = page.get_by_text("No Rows to Display", exact=True)
                    if not visible_rows_now or no_rows_message.count():
                        message = "Source account disappeared before selection; likely already merged"
                        ledger.record(pair, "skipped", message)
                        excel_status.write(pair, "skipped", message)
                        print(f"SKIP {pair.merge_from} -> {pair.merge_into}: already merged/no longer available")
                        processed += 1
                        continue
                    raise
                merge_menu = page.get_by_role("menuitem", name="Merge Accounts", exact=True)
                try:
                    merge_menu.wait_for(timeout=2_000)
                except PlaywrightTimeoutError:
                    cell.click()
                    page.keyboard.press("Shift+F10")
                    try:
                        merge_menu.wait_for(timeout=2_000)
                    except PlaywrightTimeoutError:
                        dots = row.locator("button:visible").first
                        if not dots.count():
                            raise RuntimeError("Could not open the row actions menu")
                        dots.click()
                        merge_menu.wait_for(timeout=5_000)
                merge_menu.click()
                merge_dialog = page.get_by_role("dialog", name=re.compile("Merge Accounts", re.I)).last
                merge_dialog.wait_for(timeout=15_000)
                # Momentus has rendered this lookup in multiple ways across runs.
                # Wait for the control and try labeled textbox, combobox, and plain-input variants.
                target, target_strategy = find_merge_into_field(page, merge_dialog, timeout_ms=20_000)
                print(f"MERGE INTO field found via: {target_strategy}")
                target.click()
                target.press("Control+A")
                target.type(pair.merge_into, delay=20)
                page.wait_for_timeout(800)
                exact_target, exact_strategy = find_exact_merge_target(
                    page, merge_dialog, pair.merge_into, timeout_ms=15_000
                )
                print(f"MERGE INTO result found via: {exact_strategy}")
                exact_target.click()
                merge_dialog.get_by_role("button", name="Save", exact=True).click()
                confirm = page.get_by_role("dialog", name=re.compile("Merge Accounts", re.I)).last
                confirm.get_by_role("button", name="OK", exact=True).click()
                page.get_by_text("Success", exact=True).wait_for(timeout=30_000)
                ledger.record(pair, "merged")
                excel_status.write(pair, "merged", f"Merged into {pair.merge_into}")
                print(f"MERGED {pair.merge_from} -> {pair.merge_into}")
                processed += 1
            except KeyboardInterrupt:
                raise
            except Exception as exc:
                error_message = str(exc)
                if error_message.startswith("Exact Merge Into account was not found in autocomplete:"):
                    message = f"Human Needed: Merge Into account was not found in autocomplete ({pair.merge_into})"
                    ledger.record(pair, "human_needed", message)
                    excel_status.write(pair, "human_needed", message)
                    print(f"HUMAN NEEDED {pair.merge_from} -> {pair.merge_into}: autocomplete target not found")
                    processed += 1
                    continue

                # Other errors still stop the batch for review.
                try:
                    page.screenshot(path=str(run_dir / f"error_{pair.merge_from}.png"), full_page=True)
                    (run_dir / f"error_{pair.merge_from}.html").write_text(page.content(), encoding="utf-8")
                    (run_dir / f"error_{pair.merge_from}.txt").write_text(
                        page.locator("body").inner_text(), encoding="utf-8"
                    )
                except Exception as diagnostic_exc:
                    print(f"Could not save error diagnostics: {diagnostic_exc}", file=sys.stderr)
                ledger.record(pair, "error", error_message)
                excel_status.write(pair, "error", error_message)
                print(f"ERROR {pair.merge_from} -> {pair.merge_into}: {error_message}", file=sys.stderr)
                break
            finally:
                ledger.export_csv(run_dir / "progress.csv")
                write_summary(run_dir / "summary.json", pairs, ledger, args.input, args.dry_run)
        context.close()
    excel_status.close()
    ledger.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
